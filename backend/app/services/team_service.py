"""Team management business logic."""
from typing import List, Optional, Dict
from datetime import date, timedelta
from sqlalchemy import select, func, case, literal_column, or_, delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy.sql import func as sql_func
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
from io import BytesIO

from app.models.team import (
    Person, Group, Capability, CapabilityDimension,
    Responsibility, KeyFigure
)
from app.models.project import Task, TaskCompletion
from app.schemas.team import (
    PersonCreate, PersonUpdate, PersonResponse,
    CapabilityDimensionCreate, CapabilityDimensionUpdate,
    GroupCreate, GroupUpdate, GroupResponse,
    ResponsibilityCreate, ResponsibilityUpdate, ResponsibilityResponse,
    CapabilityCreate, CapabilityUpdate,
    KeyFigureCreate,
    WorkloadPersonResponse, WorkloadGroupResponse, WorkloadMonthlySummary
)
from app.utils.helpers import (
    calculate_workload, determine_completion_status,
    check_task_conflict, find_longest_path
)


class TeamService:
    """Team management service."""

    # Capability Dimensions
    async def get_capability_dimensions(
        self, db: AsyncSession, is_admin: bool
    ) -> List[CapabilityDimension]:
        """Get all capability dimensions."""
        result = await db.execute(
            select(CapabilityDimension)
            .order_by(CapabilityDimension.is_active.desc())
        )
        return result.scalars().all()

    async def create_capability_dimension(
        self, db: AsyncSession, schema: CapabilityDimensionCreate
    ) -> CapabilityDimension:
        """Create a capability dimension."""
        db_dim = CapabilityDimension(**schema.model_dump())
        db.add(db_dim)
        await db.commit()
        await db.refresh(db_dim)
        return db_dim

    async def update_capability_dimension(
        self, db: AsyncSession, id: int,
        schema: CapabilityDimensionUpdate
    ) -> CapabilityDimension:
        """Update a capability dimension."""
        db_dim = await db.get(CapabilityDimension, id)
        if not db_dim:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"能力维度 {id} 不存在"
            )
        update_data = schema.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_dim, field, value)
        await db.commit()
        await db.refresh(db_dim)
        return db_dim

    async def delete_capability_dimension(
        self, db: AsyncSession, id: int
    ) -> bool:
        """Delete a capability dimension."""
        db_dim = await db.get(CapabilityDimension, id)
        if not db_dim:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"能力维度 {id} 不存在"
            )
        await db.delete(db_dim)
        await db.commit()
        return True

    # Persons
    async def get_persons(
        self, db: AsyncSession,
        skip: int = 0,
        limit: int = 100,
        search: Optional[str] = None
    ) -> List[Person]:
        """Get persons with pagination and search."""
        query = select(Person)

        if search:
            query = query.where(
                or_(
                    Person.name.ilike(f"%{search}%"),
                    Person.emp_id.ilike(f"%{search}%"),
                    Person.email.ilike(f"%{search}%")
                )
            )

        result = await db.execute(
            query
            .options(selectinload(Person.group))
            .order_by(Person.name)
            .offset(skip)
            .limit(limit)
        )
        return result.scalars().all()

    async def get_person(
        self, db: AsyncSession, id: int
    ) -> Optional[Person]:
        """Get person by ID."""
        result = await db.execute(
            select(Person)
            .options(
                selectinload(Person.group),
                selectinload(Person.capabilities),
                selectinload(Person.responsibilities_owner),
                selectinload(Person.responsibilities_backup),
                selectinload(Person.dev_tasks).selectinload(Task.iteration)
            )
            .where(Person.id == id)
        )
        return result.scalar_one_or_none()

    async def create_person(
        self, db: AsyncSession, schema: PersonCreate
    ) -> Person:
        """Create a person."""
        # Check if emp_id already exists
        existing = await db.execute(
            select(Person).where(Person.emp_id == schema.emp_id)
        )
        if existing.scalar_one_or_none():
            from fastapi import HTTPException
            raise HTTPException(
                status_code=400,
                detail=f"工号 {schema.emp_id} 已存在"
            )

        db_person = Person(**schema.model_dump())
        db.add(db_person)
        await db.commit()
        await db.refresh(db_person)
        return db_person

    async def update_person(
        self, db: AsyncSession, id: int,
        schema: PersonUpdate
    ) -> Person:
        """Update a person."""
        db_person = await db.get(Person, id)
        if not db_person:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"人员 {id} 不存在"
            )

        update_data = schema.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_person, field, value)

        await db.commit()
        await db.refresh(db_person)
        return db_person

    async def delete_person(
        self, db: AsyncSession, id: int
    ) -> bool:
        """Delete a person."""
        db_person = await db.get(Person, id)
        if not db_person:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"人员 {id} 不存在"
            )

        # Delete related data will cascade delete
        await db.delete(db_person)
        await db.commit()
        return True

    # Capabilities
    async def get_person_capabilities(
        self, db: AsyncSession, person_id: int
    ) -> List[Dict]:
        """Get person capabilities."""
        result = await db.execute(
            select(Capability)
            .where(Capability.person_id == person_id)
            .order_by(Capability.dimension)
        )
        capabilities = result.scalars().all()
        return [
            {
                "dimension": cap.dimension,
                "level": cap.level,
                "description": cap.description
            }
            for cap in capabilities
        ]

    async def update_person_capabilities(
        self, db: AsyncSession,
        person_id: int,
        capabilities: List[CapabilityCreate]
    ) -> List[Capability]:
        """Update person capabilities (Admin only)."""
        # Get person
        person = await db.get(Person, person_id)
        if not person:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"人员 {person_id} 不存在"
            )

        # Remove old capabilities
        await db.execute(
            delete(Capability).where(Capability.person_id == person_id)
        )

        # Create new capabilities
        for cap in capabilities:
            db_cap = Capability(**cap.model_dump())
            db.add(db_cap)

        await db.commit()
        await db.refresh(person)
        result = await db.execute(
            select(Capability).where(Capability.person_id == person_id)
        )
        return result.scalars().all()

    # Groups
    async def get_groups(
        self, db: AsyncSession,
        skip: int = 0,
        limit: int = 100
    ) -> List[Group]:
        """Get groups with pagination."""
        result = await db.execute(
            select(Group)
            .options(selectinload(Group.leader))
            .order_by(Group.name)
            .offset(skip)
            .limit(limit)
        )
        return result.scalars().all()

    async def get_group(
        self, db: AsyncSession, id: int
    ) -> Optional[Group]:
        """Get group by ID."""
        result = await db.execute(
            select(Group)
            .options(
                selectinload(Group.leader),
                selectinload(Group.members)
            )
            .where(Group.id == id)
        )
        return result.scalar_one_or_none()

    async def create_group(
        self, db: AsyncSession, schema: GroupCreate
    ) -> Group:
        """Create a group."""
        # Check if name already exists
        existing = await db.execute(
            select(Group).where(Group.name == schema.name)
        )
        if existing.scalar_one_or_none():
            from fastapi import HTTPException
            raise HTTPException(
                status_code=400,
                detail=f"小组 {schema.name} 已存在"
            )

        db_group = Group(**schema.model_dump())
        db.add(db_group)
        await db.commit()
        await db.refresh(db_group)
        return db_group

    async def update_group(
        self, db: AsyncSession, id: int,
        schema: GroupUpdate
    ) -> Group:
        """Update a group."""
        db_group = await db.get(Group, id)
        if not db_group:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"小组 {id} 不存在"
            )

        update_data = schema.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_group, field, value)

        await db.commit()
        await db.refresh(db_group)
        return db_group

    async def delete_group(
        self, db: AsyncSession, id: int
    ) -> bool:
        """Delete a group."""
        db_group = await db.get(Group, id)
        if not db_group:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"小组 {id} 不存在"
            )

        await db.delete(db_group)
        await db.commit()
        return True

    # Responsibilities
    async def get_responsibilities(
        self, db: AsyncSession,
        group_id: Optional[int] = None,
        skip: int = 0,
        limit: int = 100
    ) -> List[Responsibility]:
        """Get responsibilities with optional group filter."""
        query = select(Responsibility)

        if group_id:
            query = query.where(Responsibility.group_id == group_id)

        result = await db.execute(
            query
            .options(
                selectinload(Responsibility.owner),
                selectinload(Responsibility.backup),
                selectinload(Responsibility.group)
            )
            .order_by(Responsibility.name)
            .offset(skip)
            .limit(limit)
        )
        return result.scalars().all()

    async def create_responsibility(
        self, db: AsyncSession, schema: ResponsibilityCreate
    ) -> Responsibility:
        """Create a responsibility."""
        db_resp = Responsibility(**schema.model_dump())
        db.add(db_resp)
        await db.commit()
        await db.refresh(db_resp)
        return db_resp

    async def update_responsibility(
        self, db: AsyncSession, id: int,
        schema: ResponsibilityUpdate
    ) -> Responsibility:
        """Update a responsibility."""
        db_resp = await db.get(Responsibility, id)
        if not db_resp:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"责任田 {id} 不存在"
            )

        update_data = schema.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_resp, field, value)

        await db.commit()
        await db.refresh(db_resp)
        return db_resp

    async def delete_responsibility(
        self, db: AsyncSession, id: int
    ) -> bool:
        """Delete a responsibility."""
        db_resp = await db.get(Responsibility, id)
        if not db_resp:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"责任田 {id} 不存在"
            )

        await db.delete(db_resp)
        await db.commit()
        return True

    # Key Figures
    async def get_group_key_figures(
        self, db: AsyncSession, group_id: int
    ) -> List[KeyFigure]:
        """Get all key figures for a group."""
        result = await db.execute(
            select(KeyFigure)
            .options(selectinload(KeyFigure.person))
            .where(KeyFigure.group_id == group_id)
        )
        return result.scalars().all()

    async def create_key_figure(
        self, db: AsyncSession, group_id: int, schema: KeyFigureCreate
    ) -> KeyFigure:
        """Create a key figure for a group."""
        # Verify group exists
        group = await db.get(Group, group_id)
        if not group:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"小组 {group_id} 不存在"
            )

        # Verify person exists
        person = await db.get(Person, schema.person_id)
        if not person:
            raise HTTPException(
                status_code=404,
                detail=f"人员 {schema.person_id} 不存在"
            )

        db_key_figure = KeyFigure(
            group_id=group_id,
            type=schema.type,
            person_id=schema.person_id
        )
        db.add(db_key_figure)
        await db.commit()
        await db.refresh(db_key_figure)
        return db_key_figure

    async def delete_key_figure(
        self, db: AsyncSession, group_id: int, kf_id: int
    ) -> bool:
        """Delete a key figure."""
        db_kf = await db.get(KeyFigure, kf_id)
        if not db_kf or db_kf.group_id != group_id:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"关键人物 {kf_id} 不存在"
            )

        await db.delete(db_kf)
        await db.commit()
        return True

    # Team Structure Graph
    async def get_team_structure_graph(
        self, db: AsyncSession
    ) -> Dict:
        """Get team structure graph data for visualization."""
        # Get all groups with members
        result = await db.execute(
            select(Group)
            .options(
                selectinload(Group.leader),
                selectinload(Group.members),
                selectinload(Group.responsibilities)
                .selectinload(Responsibility.owner)
                .selectinload(Responsibility.backup),
                selectinload(Group.key_figures).selectinload(KeyFigure.person)
            )
        )
        groups = result.scalars().all()

        nodes = []
        edges = []

        # Add group nodes
        for group in groups:
            nodes.append({
                "id": f"group_{group.id}",
                "name": group.name,
                "type": "group",
                "leader_id": group.leader_id,
                "leader_name": group.leader.name if group.leader else None
            })

            # Add member nodes
            for member in group.members:
                nodes.append({
                    "id": f"person_{member.id}",
                    "name": member.name,
                    "type": "person",
                    "emp_id": member.emp_id,
                    "role": member.role
                })
                # Add edge: person belongs to group
                edges.append({
                    "source": f"person_{member.id}",
                    "target": f"group_{group.id}",
                    "type": "belongs_to"
                })

            # Add key figure edges
            for kf in group.key_figures:
                edges.append({
                    "source": f"person_{kf.person_id}",
                    "target": f"group_{group.id}",
                    "type": "key_figure",
                    "label": kf.type
                })

            # Add responsibility nodes and edges
            for resp in group.responsibilities:
                nodes.append({
                    "id": f"resp_{resp.id}",
                    "name": resp.name,
                    "type": "responsibility"
                })
                # Add edge: responsibility belongs to group
                edges.append({
                    "source": f"resp_{resp.id}",
                    "target": f"group_{group.id}",
                    "type": "belongs_to"
                })
                # Add edge: person owns responsibility
                if resp.owner:
                    edges.append({
                        "source": f"person_{resp.owner_id}",
                        "target": f"resp_{resp.id}",
                        "type": "owns"
                    })
                # Add edge: person is backup for responsibility
                if resp.backup:
                    edges.append({
                        "source": f"person_{resp.backup_id}",
                        "target": f"resp_{resp.id}",
                        "type": "backup"
                    })

        return {"nodes": nodes, "edges": edges}

    # Capability Radar
    async def get_person_capability_radar(
        self, db: AsyncSession, person_id: int
    ) -> Dict:
        """Get person capability data for radar chart."""
        result = await db.execute(
            select(Capability, CapabilityDimension)
            .join(CapabilityDimension, Capability.dimension == CapabilityDimension.name)
            .where(Capability.person_id == person_id)
            .order_by(CapabilityDimension.id)
        )
        capabilities = result.all()

        # Get all active dimensions for this person
        person = await db.get(Person, person_id)
        if not person:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"人员 {person_id} 不存在"
            )

        dimensions = []
        values = []

        for cap, dim in capabilities:
            if dim.is_active:
                dimensions.append(dim.name)
                values.append(cap.level)

        return {
            "person_id": person_id,
            "person_name": person.name,
            "dimensions": dimensions,
            "values": values
        }

    async def get_group_capability_radar(
        self, db: AsyncSession, group_id: int
    ) -> Dict:
        """Get group capability data for radar chart (average)."""
        # Get all persons in group
        result = await db.execute(
            select(Person).where(Person.group_id == group_id)
        )
        persons = result.scalars().all()

        if not persons:
            return {
                "group_id": group_id,
                "group_name": "",
                "dimensions": [],
                "values": []
            }

        # Get all capabilities for group members
        result = await db.execute(
            select(Capability, CapabilityDimension, Person.name)
            .join(CapabilityDimension, Capability.dimension == CapabilityDimension.name)
            .join(Person, Capability.person_id == Person.id)
            .where(Person.group_id == group_id, CapabilityDimension.is_active == True)
            .order_by(CapabilityDimension.id)
        )
        capabilities = result.all()

        # Calculate average per dimension
        dimension_scores = {}

        for cap, dim, person_name in capabilities:
            if dim.name not in dimension_scores:
                dimension_scores[dim.name] = {"total": 0, "count": 0}
            dimension_scores[dim.name]["total"] += cap.level
            dimension_scores[dim.name]["count"] += 1

        dimensions = list(dimension_scores.keys())
        values = [
            round(dimension_scores[dim]["total"] / dimension_scores[dim]["count"], 2)
            for dim in dimensions
        ]

        # Get group name
        group = await db.get(Group, group_id)

        return {
            "group_id": group_id,
            "group_name": group.name if group else "",
            "member_count": len(persons),
            "dimensions": dimensions,
            "values": values
        }

    # Workload Analysis
    async def get_person_workload(
        self, db: AsyncSession,
        person_id: int,
        start_date: date,
        end_date: date
    ) -> WorkloadPersonResponse:
        """Calculate person workload for a date range."""
        # Get person with tasks
        result = await db.execute(
            select(Person)
            .where(Person.id == person_id)
            .options(
                selectinload(Person.group),
                selectinload(Person.dev_tasks).selectinload(Task.iteration)
            )
        )

        person = result.scalar_one_or_none()
        if not person:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"人员 {person_id} 不存在"
            )

        tasks = person.dev_tasks

        # Calculate workload
        task_data = [
            {
                "start_date": task.start_date,
                "end_date": task.end_date,
                "man_months": task.man_months,
            }
            for task in tasks
        ]

        workload = calculate_workload(task_data, start_date, end_date)

        # Calculate task count
        task_count = len(task_data)

        return WorkloadPersonResponse(
            person_id=person_id,
            person_name=person.name,
            workload=workload,
            task_count=task_count,
            tasks=[{"id": t.id, "name": t.name} for t in tasks]
        )

    async def get_group_workload(
        self, db: AsyncSession,
        group_id: int,
        start_date: date,
        end_date: date
    ) -> WorkloadGroupResponse:
        """Calculate group workload for a date range."""
        # Get all persons in the group
        result = await db.execute(
            select(Person)
            .where(Person.group_id == group_id)
            .options(
                selectinload(Person.group),
                selectinload(Person.dev_tasks).selectinload(Task.iteration)
            )
        )

        persons = result.scalars().all()
        if not persons:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=404,
                detail=f"小组 {group_id} 不存在"
            )

        # Calculate workload for each person
        member_workloads = []
        group_workload = 0.0

        # Calculate workloads for all members
        for person in persons:
            task_data = [
                {
                    "start_date": task.start_date,
                    "end_date": task.end_date,
                    "man_months": task.man_months,
                }
                for task in person.dev_tasks
            ]
            workload = calculate_workload(task_data, start_date, end_date)
            group_workload += workload

            member_workloads.append(
                WorkloadPersonResponse(
                    person_id=person.id,
                    person_name=person.name,
                    workload=workload,
                    task_count=len(person.dev_tasks),
                    tasks=[{"id": t.id, "name": t.name} for t in person.dev_tasks]
                )
            )

        # Sort by workload descending
        member_workloads.sort(key=lambda x: x.workload, reverse=True)

        # Get group name
        result = await db.execute(select(Group).where(Group.id == group_id))
        group = result.scalar_one()

        return WorkloadGroupResponse(
            group_id=group_id,
            group_name=group.name if group else "",
            workload=group_workload,
            member_count=len(persons),
            member_workloads=member_workloads,
        )

    async def get_monthly_workload_summary(
        self, db: AsyncSession,
        month: int,
        year: int
    ) -> WorkloadMonthlySummary:
        """Get monthly workload summary."""
        # Calculate month start and end dates
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)

        # Get all persons with their tasks in the month
        result = await db.execute(
            select(Person, Task)
            .join(Task, Task.developer_id == Person.id)
            .where(
                Task.start_date >= start_date,
                Task.end_date <= end_date
            )
        )

        total_workload = 0.0
        person_workloads: List[Dict] = []

        for person, task in result.all():
            task_duration_days = (task.end_date - task.start_date).days + 1
            daily_workload = task.man_months / 30  # 假设每月30天
            person_workload = daily_workload * task_duration_days

            total_workload += person_workload

            # Track per-person workload
            person_workloads.append({
                "person_id": person.id,
                "person_name": person.name,
                "workload": person_workload,
                "task_count": 1,
            })

        # Calculate average workload
        avg_workload = total_workload / len(person_workloads) if person_workloads else 0

        # Find highest load person
        highest_load_person = None
        highest_load = 0.0

        for pwl in person_workloads:
            if pwl["workload"] > highest_load:
                highest_load = pwl["workload"]
                highest_load_person = pwl["person_name"]

        return WorkloadMonthlySummary(
            total_workload=round(total_workload, 2),
            avg_workload=round(avg_workload, 2),
            highest_load=highest_load_person,
            highest_load_value=round(highest_load, 2) if highest_load_person else None,
            person_workloads=person_workloads,
        )


# Export
async def export_achievement_to_excel(
    db: AsyncSession,
    version_ids: Optional[List[int]] = None,
    iteration_ids: Optional[List[int]] = None,
    person_ids: Optional[List[int]] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    """Export achievement data to Excel."""

    # Build query conditions
    base_query = select(
        Task, TaskCompletion, Person
    ).select_from(
        Task
    ).join(
        TaskCompletion, Task.id == TaskCompletion.task_id
    ).join(
        Person, Task.developer_id == Person.id
    )

    # Apply filters
    if version_ids:
        base_query = base_query.where(Task.version_id.in_(version_ids))

    if iteration_ids:
        base_query = base_query.where(Task.iteration_id.in_(iteration_ids))

    if person_ids:
        base_query = base_query.where(Task.developer_id.in_(person_ids))

    if start_date:
        base_query = base_query.where(Task.end_date >= start_date)

    if end_date:
        base_query = base_query.where(Task.start_date <= end_date)

    # Execute query
    result = await db.execute(base_query)
    rows = result.all()

    # Prepare data for Excel export
    excel_data = []
    for task, completion, person in rows:
        if completion:
            completion_date = completion.actual_end_date
            completion_status = determine_completion_status(task.end_date, completion_date)
        else:
            completion_date = None
            completion_status = None

        excel_data.append({
            "任务ID": task.id,
            "任务名称": task.name,
            "工号": person.emp_id,
            "姓名": person.name,
            "计划开始": task.start_date.strftime("%Y-%m-%d") if task.start_date else "",
            "计划结束": task.end_date.strftime("%Y-%m-%d") if task.end_date else "",
            "实际完成": completion_date.strftime("%Y-%m-%d %H:%M:%S") if completion_date else "",
            "完成状态": completion_status,
            "提前完成": 1 if completion_status == "early" else 0,
            "准时完成": 1 if completion_status == "on_time" else 0,
            "略微超期": 1 if completion_status == "slightly_late" else 0,
            "严重超期": 1 if completion_status == "severely_late" else 0,
        })

    # Create workbook
    wb = Workbook()
    ws = wb.active

    # Add summary sheet
    # Header row
    ws.append([
        "任务ID",
        "任务名称",
        "工号",
        "姓名",
        "计划开始",
        "计划结束",
        "实际完成",
        "完成状态",
        "提前完成",
        "准时完成",
        "略微超期",
        "严重超期",
    ])

    ws.title = "任务达成统计"
    header_font = Font(name='微软雅黑', size=11, bold=True, color='#FFFFFF')

    # Data rows
    for row in excel_data:
        ws.append([
            row["任务ID"],
            row["任务名称"],
            row["工号"],
            row["姓名"],
            row["计划开始"],
            row["计划结束"],
            row["实际完成"],
            row["完成状态"],
            row["提前完成"],
            row["准时完成"],
            row["略微超期"],
            row["严重超期"],
        ])

    # Auto-adjust column widths
    column_widths = {
        "A": 10,  # 任务ID
        "B": 30,  # 任务名称
        "C": 12,  # 工号
        "D": 12,  # 姓名
        "E": 15,  # 计划开始
        "F": 15,  # 计划结束
        "G": 20,  # 实际完成
        "H": 12,  # 完成状态
        "I": 10,  # 提前完成
        "J": 10,  # 准时完成
        "K": 10,  # 略微超期
        "L": 10,  # 严重超期
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Style the header
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='#FFFFFF')
        cell.alignment = Alignment(horizontal="center")

    # Style rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column_letter == "H":  # 完成状态列
                status = cell.value
                if status == "early":
                    cell.fill = PatternFill(start_color="90EE90", fill_type="solid")
                elif status == "on_time":
                    cell.fill = PatternFill(start_color="00FF00", fill_type="solid")
                elif status == "slightly_late":
                    cell.fill = PatternFill(start_color="FFEB3B", fill_type="solid")
                elif status == "severely_late":
                    cell.fill = PatternFill(start_color="F5222D", fill_type="solid")

            cell.font = Font(name="微软雅黑", size=10, color="#000000")
            cell.alignment = Alignment(horizontal="center")

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# Global service instance
team_service = TeamService()
